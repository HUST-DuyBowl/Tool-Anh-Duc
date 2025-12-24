# filechinh.py
# v22.0 - Sửa lỗi hyperlink triệt để bằng cách tự định nghĩa Font.

import os
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font # << Quan trọng: Import Font để định nghĩa style
import img2pdf
from pdf2docx import Converter
from spire.doc import Document, FileFormat
from datetime import datetime
from pathlib import Path

def analyze_and_sort_folder(folder_path):
    """
    Phân tích và phân loại các file, hỗ trợ cả .doc và .zip.
    """
    files_to_skip = []
    archive_files = []
    doc_files = defaultdict(list)
    image_files = []

    zalo_pattern = re.compile(r'^z\d+_[a-f0-9]{32}\.(jpg|jpeg|png)$', re.IGNORECASE)
    image_group_pattern = re.compile(r'^(.*?)([\s_-]*tờ[\s_-]*\d+)(\..*)$', re.IGNORECASE)
    
    for filename in os.listdir(folder_path):
        if not os.path.isfile(os.path.join(folder_path, filename)):
            continue

        if zalo_pattern.match(filename):
            files_to_skip.append(filename)
            continue
            
        base_name, extension = os.path.splitext(filename)
        extension = extension.lower()

        if extension in ['.rar', '.zip']:
            archive_files.append(filename)
        elif extension in ['.pdf', '.docx', '.doc']:
            doc_files[base_name].append(extension)
        elif extension in ['.jpg', '.jpeg', '.png']:
            image_files.append(filename)
        else:
            files_to_skip.append(filename)

    pdf_doc_pairs = []
    pdf_only = []
    doc_only = []

    for base_name, exts in doc_files.items():
        has_pdf = '.pdf' in exts
        doc_ext = '.docx' if '.docx' in exts else ('.doc' if '.doc' in exts else None)

        if has_pdf and doc_ext:
            pdf_doc_pairs.append((base_name, doc_ext))
        elif has_pdf:
            pdf_only.append(base_name)
        elif doc_ext:
            doc_only.append((base_name, doc_ext))
            
    potential_groups = defaultdict(list)
    single_images = []
    for img_filename in image_files:
        match = image_group_pattern.match(img_filename)
        if match:
            group_name = match.group(1).strip(' _-').lower()
            potential_groups[group_name].append(img_filename)
        else:
            single_images.append(img_filename)

    final_image_groups = {}
    for group_name, files in potential_groups.items():
        if len(files) > 1:
            final_image_groups[group_name] = sorted(files)
        else:
            single_images.extend(files)

    return (
        sorted(pdf_doc_pairs),
        sorted(pdf_only),
        sorted(doc_only),
        sorted(archive_files),
        final_image_groups,
        sorted(single_images),
        sorted(files_to_skip)
    )

def append_row_to_xlsx(xlsx_path, data_row):
    """
    Hàm ghi Excel hoàn thiện, tạo hyperlink chính xác bằng cách tự định nghĩa style.
    """
    header = [
        "STT", "Tên file", "File PDF", "File DOCX/DOC", "File Nén", "File Khác", 
        "Năm học", "Phân loại", "Lớp", "Chương học", "Bài học"
    ]
    path_indices = [2, 3, 4]

    workbook = None
    if not os.path.exists(xlsx_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "FileLog"
        sheet.append(header)
    else:
        try:
            workbook = openpyxl.load_workbook(xlsx_path)
        except Exception:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = xlsx_path.replace('.xlsx', f'_loi_{timestamp}.xlsx')
            os.rename(xlsx_path, backup_path)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "FileLog"
            sheet.append(header)

    if workbook:
        sheet = workbook.active
        new_row_num = sheet.max_row + 1

        # *** LOGIC SỬA LỖI: Tự định nghĩa style cho hyperlink ***
        hyperlink_font = Font(color="0563C1", underline="single")

        for col_num, cell_value in enumerate(data_row, 1):
            cell = sheet.cell(row=new_row_num, column=col_num)
            
            if col_num - 1 in path_indices and cell_value and os.path.exists(str(cell_value)):
                full_path = str(cell_value)
                file_name = os.path.basename(full_path)
                uri = Path(os.path.abspath(full_path)).as_uri()

                cell.value = file_name
                cell.hyperlink = uri
                cell.font = hyperlink_font # Áp dụng style tự định nghĩa
            else:
                cell.value = cell_value

        try:
            workbook.save(xlsx_path)
        except PermissionError:
             print(f"Lỗi: Không có quyền ghi vào file {xlsx_path}. File có thể đang được mở.")

def taopdftudocx(base_name, pathdoc):
    temp_pdf_path = os.path.join(os.path.dirname(pathdoc), f"{base_name}_temp.pdf")
    try:
        document = Document()
        document.LoadFromFile(pathdoc)
        document.SaveToFile(temp_pdf_path, FileFormat.PDF)
        document.Dispose()
        document.Close()
        return temp_pdf_path
    except Exception as e:
        print(f"Lỗi khi chuyển đổi DOC/DOCX sang PDF cho '{base_name}': {e}")
        return None

def taofile_from_images(base_name, listfileanh, output_dir):
    pdf_path = os.path.join(output_dir, f'{base_name}.pdf')
    doc_path = os.path.join(output_dir, f'{base_name}.docx')
    try:
        with open(pdf_path, "wb") as f:
            f.write(img2pdf.convert(listfileanh))
        
        cv = Converter(pdf_path)
        cv.convert(doc_path)
        cv.close()
        return pdf_path, doc_path
    except Exception as e:
        print(f"Lỗi khi tạo file từ ảnh cho '{base_name}': {e}")
        return None, None